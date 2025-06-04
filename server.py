import os
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from playwright.sync_api import sync_playwright
import time

# Constants
url = "https://www.marutisuzuki.com/genuine-accessories/grand-vitara-accessories"
PAGE_LIMIT = 33

part_numbers = []
part_names = []
mrps = []
image_urls = []


def scrape_pages():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto(url)

        for current_page in range(1, PAGE_LIMIT + 1):
            print(f"Scraping page {current_page}...")
            try:
                page.wait_for_selector('.sliderBox')
                soup = BeautifulSoup(page.content(), 'html.parser')
                product_cards = soup.find_all('div', class_='sliderBox')

                for card in product_cards:
                    try:
                        part_number = card.get('data-partno', '').strip()
                        part_name = card.get('data-partname', '').strip()
                        mrp = card.get('data-price', '').strip()
                        image_src = card.find('img')['src']
                        image_url = image_src if image_src.startswith(
                            'http') else f"https://www.marutisuzuki.com{image_src}"

                        if not part_number or not part_name or not mrp:
                            print("Missing essential product info, skipping...")
                            continue

                        part_numbers.append(part_number)
                        part_names.append(part_name)
                        mrps.append(mrp)
                        image_urls.append(image_url)

                    except Exception as e:
                        print(f"Error parsing product card: {e}")
                        continue

                if current_page < PAGE_LIMIT:
                    next_button = page.query_selector(
                        "//a[contains(text(), 'Next')]")
                    if next_button:
                        next_button.click()
                        print("Waiting for the next page to load...")
                        page.wait_for_timeout(3000)
                    else:
                        print(
                            f"'Next' button not found on page {current_page}")
                        break
            except Exception as e:
                print(f"Error scraping page {current_page}: {e}")
                continue

        browser.close()


scrape_pages()

# Create image directory
if not os.path.exists('images'):
    os.makedirs('images')

image_paths = []
for idx, img_url in enumerate(image_urls):
    try:
        print(f"Downloading image {idx+1}: {img_url}")
        img_data = requests.get(img_url, timeout=10).content
        time.sleep(1)
        img_filename = f"images/part_image_{idx}.jpg"
        with open(img_filename, 'wb') as img_file:
            img_file.write(img_data)
        image_paths.append(img_filename)
    except Exception as e:
        print(f"Error downloading image {img_url}: {e}")
        image_paths.append(None)

# Save to Excel
data = {
    'Part Number': part_numbers,
    'Part Name': part_names,
    'MRP': mrps,
    'Image': image_paths
}
df = pd.DataFrame(data)

wb = Workbook()
ws = wb.active
headers = ['Part Number', 'Part Name', 'MRP', 'Image']
ws.append(headers)

for col in ['A', 'B', 'C', 'D']:
    ws.column_dimensions[col].width = 30

for i, row in df.iterrows():
    try:
        ws.cell(row=i+2, column=1, value=row['Part Number'])
        ws.cell(row=i+2, column=2, value=row['Part Name'])
        ws.cell(row=i+2, column=3, value=row['MRP'])
        if row['Image']:
            img = ExcelImage(row['Image'])
            img.height = 100
            img.width = 100
            ws.row_dimensions[i+2].height = 100
            ws.add_image(img, f"D{i+2}")
        ws.cell(
            row=i+2, column=4).alignment = Alignment(horizontal='center', vertical='center')
    except Exception as e:
        print(f"Error adding data to Excel for row {i+1}: {e}")
        continue

try:
    wb.save('scraped_parts_with_images_playwright.xlsx')
    print("✅ Excel file saved with images!")
except Exception as e:
    print(f"❌ Error saving Excel file: {e}")
